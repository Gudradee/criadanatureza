
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C_TEAL    = "38A3A5"
C_TEAL_FG = "FFFFFF"
C_HEADER  = "1E3A4C"
C_SUBTLE  = "F1F5F9"
C_TOTAL   = "E2E8F0"

_BORDER_THIN = Side(style="thin", color="CBD5E1")
BORDER_ALL   = Border(left=_BORDER_THIN, right=_BORDER_THIN, top=_BORDER_THIN, bottom=_BORDER_THIN)

FONT_TITLE  = Font(name="Calibri", size=14, bold=True, color=C_HEADER)
FONT_SUB    = Font(name="Calibri", size=10, color="475569")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=C_TEAL_FG)
FONT_TOTAL  = Font(name="Calibri", size=11, bold=True, color=C_HEADER)
FONT_CELL   = Font(name="Calibri", size=11, color="0F172A")

FILL_HEADER = PatternFill("solid", fgColor=C_TEAL)
FILL_TOTAL  = PatternFill("solid", fgColor=C_TOTAL)
FILL_SUBTLE = PatternFill("solid", fgColor=C_SUBTLE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

FMT_BRL  = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_INT  = '#,##0'
FMT_PCT  = '0.0"%"'
FMT_DATE = 'dd/mm/yyyy'
FMT_DT   = 'dd/mm/yyyy hh:mm'

_BRT = timezone(timedelta(hours=-3))

def _to_brt(value):

    if value is None:
        return None
    try:
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(_BRT).replace(tzinfo=None)
    except Exception:
        return value

class SheetBuilder:

    def __init__(self, ws):
        self.ws = ws
        self.row = 1
        self._col_widths: dict[int, int] = {}

    def _track_width(self, col_idx: int, text):
        if text is None:
            return
        length = len(str(text))
        cur = self._col_widths.get(col_idx, 8)
        if length + 2 > cur:
            self._col_widths[col_idx] = min(length + 2, 60)

    def apply_widths(self):
        for col, w in self._col_widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w

    def title(self, text: str, subtitle: str | None = None, span: int = 8):
        cell = self.ws.cell(row=self.row, column=1, value=text)
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_LEFT
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=span)
        self.row += 1
        if subtitle:
            sub = self.ws.cell(row=self.row, column=1, value=subtitle)
            sub.font = FONT_SUB
            sub.alignment = ALIGN_LEFT
            self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=span)
            self.row += 1
        self.row += 1

    def blank(self, n: int = 1):
        self.row += n

    def header(self, columns: list[str]):
        for i, h in enumerate(columns, start=1):
            cell = self.ws.cell(row=self.row, column=i, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
            self._track_width(i, h)
        self.ws.row_dimensions[self.row].height = 22
        self.row += 1

    def row_values(self, values: list, formats: list[str | None] | None = None,
                   alignments: list[Alignment | None] | None = None,
                   is_total: bool = False):

        for i, v in enumerate(values, start=1):
            cell = self.ws.cell(row=self.row, column=i, value=v)
            cell.border = BORDER_ALL
            if is_total:
                cell.font = FONT_TOTAL
                cell.fill = FILL_TOTAL
            else:
                cell.font = FONT_CELL

            default_align = ALIGN_LEFT
            if isinstance(v, (int, float)):
                default_align = ALIGN_RIGHT
            elif isinstance(v, datetime):
                default_align = ALIGN_CENTER
            if alignments and i - 1 < len(alignments) and alignments[i - 1] is not None:
                cell.alignment = alignments[i - 1]
            else:
                cell.alignment = default_align

            if formats and i - 1 < len(formats) and formats[i - 1]:
                cell.number_format = formats[i - 1]

            disp = v
            if isinstance(v, datetime):
                disp = v.strftime("%d/%m/%Y %H:%M")
            elif isinstance(v, float):
                disp = f"R$ {v:,.2f}"
            self._track_width(i, disp)
        self.row += 1

    def table(self, headers: list[str], rows: list[list], formats: list[str | None] | None = None,
              total_row: list | None = None):

        self.header(headers)
        for r in rows:
            self.row_values(r, formats=formats)
        if total_row is not None:
            self.row_values(total_row, formats=formats, is_total=True)
        self.blank(1)

def novo_workbook(meta: dict | None = None) -> Workbook:

    wb = Workbook()

    default = wb.active
    wb.remove(default)
    if meta:
        wb.properties.title    = meta.get("title", "Export")
        wb.properties.creator  = meta.get("creator", "CDN — Cria da Natureza")
        wb.properties.subject  = meta.get("subject", "Relatório operacional")
        wb.properties.description = meta.get("description", "")
    return wb

def to_response_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def cabecalho_padrao(builder: SheetBuilder, *, titulo: str, periodo: str,
                     usuario: str | None = None, filtros: list[str] | None = None,
                     span: int = 8):

    builder.title(titulo, subtitle=periodo, span=span)
    gerado_em = datetime.now(_BRT).strftime("%d/%m/%Y %H:%M")
    info = f"Gerado em {gerado_em}"
    if usuario:
        info += f" · por {usuario}"
    cell = builder.ws.cell(row=builder.row, column=1, value=info)
    cell.font = FONT_SUB
    cell.alignment = ALIGN_LEFT
    builder.ws.merge_cells(start_row=builder.row, start_column=1, end_row=builder.row, end_column=span)
    builder.row += 1
    if filtros:
        for f in filtros:
            cell = builder.ws.cell(row=builder.row, column=1, value=f"Filtro: {f}")
            cell.font = FONT_SUB
            cell.alignment = ALIGN_LEFT
            builder.ws.merge_cells(start_row=builder.row, start_column=1, end_row=builder.row, end_column=span)
            builder.row += 1
    builder.blank(1)
